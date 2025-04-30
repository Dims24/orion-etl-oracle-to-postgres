#!/usr/bin/env python3
import sys, os
# Добавляем в поисковый путь Python корень проекта (чтобы импорт `logger` работал)
sys.path.insert(
    0,
    os.path.abspath(
        os.path.join(os.path.dirname(__file__), '..')
    )
)
import argparse
import os
import yaml
import pandas as pd
from logger import setup_logging

# Шаблоны по умолчанию для глобального конфига
GLOBAL_TEMPLATE = {
    'logging': None,
    'tables_folder': None,
    'batch_size': 5000,
    'auto_mapping_plugin': 'auto_mapping',
    'fetcher_plugin': 'default_fetcher',
    'transform_plugins': ['default_transform'],
    'validation_plugins': ['default_validation'],
    'loader_plugin': 'default_loader',
    'connectors': {
        'oracle': {'client_lib_dir': None, 'user': '', 'password': '', 'host': '', 'port': None, 'service_name': ''},
        'postgres': {'user': '', 'password': '', 'host': '', 'port': None, 'database': ''}
    },
    'table_files': []
}

# Шаблон таблицы
TABLE_TEMPLATE = {
    'source_table': None,
    'source_schema': None,
    'target_table': None,
    'target_schema': 'public',
    'fetcher_plugin': None,
    'mappings': None,
    'where': None,
    'transform_override': False,
    'transform_plugins': None,
    'loader_plugin': None
}

# Шаблон правила маппинга
MAPPING_TEMPLATE = {
    'source': None,
    'target': None,
    'transform': None,
    'plugin': None,
    'lookup': None,
    'validation': None
}


def get_str(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    return s or None


def parse_bool(val):
    if isinstance(val, bool): return val
    if isinstance(val, int): return bool(val)
    if isinstance(val, str): return val.strip().lower() in ('true', '1', 'yes', 'y')
    return False


def parse_list(val):
    if pd.isna(val): return None
    if isinstance(val, str):
        lst = [item.strip() for item in val.split(',') if item.strip()]
        return lst or None
    if isinstance(val, (list, tuple)):
        lst = [str(item).strip() for item in val if str(item).strip()]
        return lst or None
    return None


def load_table_sheets(xlsx_path, logger):
    excel = pd.ExcelFile(xlsx_path)
    sheets = {}
    # Подготовим множества обязательных колонок
    template_first = set(TABLE_TEMPLATE.keys())
    template_mapping = set(MAPPING_TEMPLATE.keys())

    for idx, sheet in enumerate(excel.sheet_names):
        df = excel.parse(sheet)
        cols_lower = [col.strip().lower() for col in df.columns]
        # Для первого листа – TABLE_TEMPLATE, для остальных – MAPPING_TEMPLATE
        expected = template_first if idx == 0 else template_mapping
        missing = expected - set(cols_lower)
        if missing:
            kind = 'основной' if idx == 0 else 'маппинга'
            logger.error(f"Лист '{sheet}' ({kind}) пропущен: отсутствуют колонки {sorted(missing)}")
            continue
        df.columns = cols_lower
        sheets[sheet] = df

    return excel, sheets


def parse_mapping_sheet(excel, sheet_name, logger):
    try:
        df = excel.parse(sheet_name)
    except Exception:
        logger.error(f"Лист маппинга '{sheet_name}' не найден")
        return None
    df.columns = [col.strip().lower() for col in df.columns]
    rules = []
    for _, row in df.iterrows():
        rule = dict(MAPPING_TEMPLATE)
        # source, target
        if get_str(row.get('source')): rule['source'] = get_str(row['source'])
        if get_str(row.get('target')): rule['target'] = get_str(row['target'])
        # transform
        tr_list = parse_list(row.get('transform'))
        rule['transform'] = tr_list or get_str(row.get('transform'))
        # plugin
        if get_str(row.get('plugin')): rule['plugin'] = get_str(row['plugin'])
        # lookup
        lookup_cell = get_str(row.get('lookup'))
        if lookup_cell:
            if ':' in lookup_cell:
                parts = [p.strip() for p in lookup_cell.split(':')]
                lookup_config = {'table': None, 'key_column': None, 'value_column': None, 'on_missing': None}
                # Берём основную часть до ':' и после '='
                base = parts[-1]
                if '=' not in base:
                    logger.error(f"Некорректный формат lookup '{lookup_cell}' в листе '{sheet_name}'")
                key_val = base.split('=', 1)
                # Парсим ключ таблицы
                table_key = key_val[0]
                if '.' in table_key:
                    table, key = table_key.split('.', 1)
                    lookup_config['table'] = table
                    lookup_config['key_column'] = key
                else:
                    logger.error(f"Некорректный формат lookup mapping '{table_key}' в листе '{sheet_name}'")
                # Парсим значение колонки, если есть
                if len(key_val) > 1:
                    right = key_val[1]
                    if '.' in right:
                        _, val_col = right.split('.', 1)
                        lookup_config['value_column'] = val_col
                    else:
                        lookup_config['value_column'] = right
                # Парсим on_missing из первой части, если не 'null'
                on_missing = parts[0] if parts and parts[0].lower() != 'null' else None
                if on_missing:
                    lookup_config['on_missing'] = on_missing
                rule['lookup'] = lookup_config
            else:
                logger.error(f"Некорректный формат lookup '{lookup_cell}' в листе '{sheet_name}'")
        # validation
        val_list = parse_list(row.get('validation'))
        if val_list:
            vrules = []
            for v in val_list:
                typ, detail = v.split(':',1)
                vr = {'type': typ}
                if typ in ('regex','range'):
                    vr['pattern'] = detail
                elif typ == 'lookup':
                    # detail: table.key or table.key:on_missing
                    parts = detail.split(':')
                    table_key = parts[0]
                    on_missing = parts[1] if len(parts) > 1 else None
                    tname, key = table_key.split('.', 1)
                    vr['lookup'] = {
                        'table': tname,
                        'key_column': key,
                        'on_missing': on_missing
                    }
                    # on_fail maps to behavior when validation fails
                    if on_missing:
                        vr['on_fail'] = on_missing
                    vrules.append(vr)
            rule['validation']=vrules
        # clean None
        rules.append({
            k: v for k, v in rule.items() if v is not None or k in ('source', 'target')
        })
    return rules


def write_yaml(path, content, logger):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        yaml.dump(content, f, default_flow_style=False, sort_keys=False, allow_unicode=True,indent= 4)
    logger.info(f"Записан файл конфигурации: {path}")


def main_generate(tables_folder: str, xlsx_file: str):
    logger = setup_logging()
    excel, sheets = load_table_sheets(xlsx_file, logger)
    if not sheets:
        logger.error(f"Нет корректных листов для обработки в '{xlsx_file}'")
        return

    # берем только первый лист как основной
    main_sheet = list(sheets.keys())[0]
    df = sheets[main_sheet]

    logger.header(f"Основной лист {main_sheet}", "")

    cfg_path = os.path.join(os.getcwd(), 'config', f"{main_sheet}.yaml")

    # если файл уже есть — загружаем его секцию 'global', иначе создаём с TEMPLATE
    if os.path.exists(cfg_path):
        with open(cfg_path, 'r', encoding='utf-8') as f:
            loaded = yaml.safe_load(f) or {}
        global_cfg = loaded.get('global', {}).copy()
    else:
        global_cfg = dict(GLOBAL_TEMPLATE)

    # всегда обновляем только эти два поля
    global_cfg['tables_folder'] = tables_folder

    # создаём папку для отдельных table-файлов
    table_dir = os.path.join(os.getcwd(), 'config', tables_folder)
    os.makedirs(table_dir, exist_ok=True)

    table_files = []
    for _, row in df.iterrows():
        src = get_str(row.get('source_table'))
        tgt = get_str(row.get('target_table'))
        if not src or not tgt:
            logger.warning(f"Строка пропущена: source_table ({src}) или target_table ({tgt}) не указаны")
            continue

        # схемы и плагины
        src_schema = get_str(row.get('source_schema')) or ''
        tgt_schema = get_str(row.get('target_schema'))
        if not tgt_schema and '.' in tgt:
            tgt_schema, tgt = tgt.split('.', 1)
        tgt_schema = tgt_schema or TABLE_TEMPLATE['target_schema']

        tbl_cfg = {
            'source_table':     src,
            'source_schema':    src_schema,
            'target_table':     tgt,
            'target_schema':    tgt_schema,
            'fetcher_plugin':   get_str(row.get('fetcher_plugin')),
            'where':            get_str(row.get('where')),
            'transform_override': parse_bool(row.get('transform_override')),
            'transform_plugins': parse_list(row.get('transform_plugins')),
            'loader_plugin':    get_str(row.get('loader_plugin'))
        }

        # маппинг из других листов
        map_sheet = get_str(row.get('mappings'))
        if map_sheet:
            tbl_cfg['mappings'] = parse_mapping_sheet(excel, map_sheet, logger)

        # убираем None и пишем файл
        tbl_cfg = {k: v for k, v in tbl_cfg.items() if v is not None}
        fname = f"{tgt}.yaml"
        path = os.path.join(table_dir, fname)
        write_yaml(path, tbl_cfg, logger)
        table_files.append(fname)

    # перезаписываем только список файлов
    global_cfg['table_files'] = table_files
    write_yaml(cfg_path, {'global': global_cfg}, logger)




from openpyxl.utils import get_column_letter

def generate_xlsx_from_yaml(tables_folder: str, xlsx_file: str):
    logger = setup_logging()
    config_root = os.path.join(os.getcwd(), 'config')
    global_files = [f for f in os.listdir(config_root)
                    if f.endswith('.yaml') and f != f"{tables_folder}.yaml"]

    with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
        for global_file in global_files:
            base_name = os.path.splitext(global_file)[0]
            safe_main = base_name if len(base_name) <= 31 else base_name[:28] + '...'

            # Загружаем глобальный конфиг
            with open(os.path.join(config_root, global_file), 'r', encoding='utf-8') as gf:
                global_cfg = yaml.safe_load(gf).get('global', {})

            # 1) Собираем info о том, у каких таблиц есть mappings
            tables = global_cfg.get('table_files', [])
            has_map = {}
            for tbl_fname in tables:
                tbl_path = os.path.join(config_root, tables_folder, tbl_fname)
                with open(tbl_path, 'r', encoding='utf-8') as tf:
                    tbl_cfg = yaml.safe_load(tf)
                # если в yaml есть ключ 'mappings' — запомним истинну и имя листа
                if tbl_cfg.get('mappings'):
                    # имя листа, на который будет ссылка
                    has_map[tbl_fname] = os.path.splitext(tbl_fname)[0]
                else:
                    has_map[tbl_fname] = None

            # 2) Делаем главный DataFrame
            main_rows = []
            for tbl_fname in tables:
                tbl_path = os.path.join(config_root, tables_folder, tbl_fname)
                with open(tbl_path, 'r', encoding='utf-8') as tf:
                    tbl_cfg = yaml.safe_load(tf)

                main_rows.append({
                    'source_table':      tbl_cfg.get('source_table'),
                    'source_schema':     tbl_cfg.get('source_schema'),
                    'target_table':      tbl_cfg.get('target_table'),
                    'target_schema':     tbl_cfg.get('target_schema'),
                    'fetcher_plugin':    tbl_cfg.get('fetcher_plugin'),
                    'mappings':          has_map[tbl_fname],   # либо имя листа, либо None
                    'where':             tbl_cfg.get('where'),
                    'transform_override': 'true' if tbl_cfg.get('transform_override') else 'false',
                    'transform_plugins':  ','.join(tbl_cfg.get('transform_plugins') or []),
                    'loader_plugin':     tbl_cfg.get('loader_plugin')
                })
            df_main = pd.DataFrame(main_rows)
            df_main.to_excel(writer, sheet_name=safe_main, index=False)

            # 3) Вешаем hyperlink только там, где has_map != None
            wb = writer.book
            ws = wb[safe_main]
            col_idx = df_main.columns.get_loc('mappings') + 1
            col_letter = get_column_letter(col_idx)
            for row_idx, tbl_fname in enumerate(tables, start=2):
                map_name = has_map[tbl_fname]
                if not map_name:
                    continue
                cell = ws[f"{col_letter}{row_idx}"]
                cell.value = map_name
                cell.hyperlink = f"#{map_name}!A1"
                cell.style = "Hyperlink"

            # 4) Генерация листов маппинга как раньше...
            for tbl_fname in tables:
                map_name = has_map[tbl_fname]
                if not map_name:
                    continue
                tbl_path = os.path.join(config_root, tables_folder, tbl_fname)
                with open(tbl_path, 'r', encoding='utf-8') as tf:
                    tbl_cfg = yaml.safe_load(tf)

                map_rows = []
                for rule in tbl_cfg.get('mappings', []):
                    # ... сбор полей transform, lookup, validation и т.д.
                    # (тот же код, что был раньше)
                    tr = rule.get('transform')
                    transform_str = ','.join(tr) if isinstance(tr, list) else tr
                    lookup_cfg = rule.get('lookup')
                    if lookup_cfg:
                        base = f"{lookup_cfg['table']}.{lookup_cfg['key_column']}"
                        on_missing = lookup_cfg.get('on_missing')
                        if on_missing is None:
                            on_missing = 'null'
                        if lookup_cfg.get('value_column'):
                            base += f"={lookup_cfg['table']}.{lookup_cfg['value_column']}"
                        lookup_str = f"{on_missing}:{base}"
                    else:
                        lookup_str = None

                    val_list = []
                    for vr in (rule.get('validation') or []):
                        vtype = vr.get('type')
                        if vtype in ('regex', 'range'):
                            val_list.append(f"{vtype}:{vr['pattern']}")
                        elif vtype == 'lookup':
                            lv = vr['lookup']
                            part = f"{lv['table']}.{lv['key_column']}"
                            if lv.get('on_missing'):
                                part += f":{lv['on_missing']}"
                            val_list.append(f"lookup:{part}")
                    validation_str = ','.join(val_list) if val_list else None

                    map_rows.append({
                        'source':     rule.get('source'),
                        'target':     rule.get('target'),
                        'transform':  transform_str,
                        'plugin':     rule.get('plugin'),
                        'lookup':     lookup_str,
                        'validation': validation_str
                    })

                df_map = pd.DataFrame(map_rows)
                safe_map = map_name if len(map_name) <= 31 else map_name[:28] + '...'
                df_map.to_excel(writer, sheet_name=safe_map, index=False)

    logger.info(f"Сгенерирован файл Excel: {xlsx_file}")


def main():
    parser = argparse.ArgumentParser(description='Генерация YAML-конфигов на основе описания в XLSX или обратная генерация XLSX из YAML.')
    parser.add_argument('--tables_folder', default='tables', help='Папка для YAML-файлов таблиц внутри config')
    parser.add_argument('--xlsx_file', default='data/main.xlsx', help='Путь к XLSX-файлу (для чтения или записи)')
    parser.add_argument('--reverse', action='store_true', help='Если передан, генерирует XLSX из YAML вместо генерации YAML')
    args = parser.parse_args()

    if args.reverse:
        generate_xlsx_from_yaml(args.tables_folder, args.xlsx_file)
    else:
        main_generate(args.tables_folder, args.xlsx_file)

if __name__ == '__main__':
    main()
